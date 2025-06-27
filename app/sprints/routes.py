from flask import request, jsonify, abort, render_template, send_file, current_app
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os

from . import sprints_bp
from .. import db
from ..models import Sprint, Task, Column, TaskStatus, Backlog
# from ..backlog.routes import serialize_task  # Removido para evitar problemas
from ..utils.serializers import serialize_task_for_sprints

# Função auxiliar local para serializar tarefas
def serialize_task(task):
    """Versão simplificada para uso local no módulo de sprints."""
    if not task:
        return None
    
    return {
        'id': task.id,
        'title': task.title or 'Sem título',
        'description': task.description or '',
        'priority': task.priority or 'Média',
        'specialist_name': task.specialist_name or 'Não atribuído',
        'estimated_effort': task.estimated_effort or 0,
        'logged_time': task.logged_time or 0,
        'position': task.position or 0,
        'status': str(task.status) if task.status else 'TODO',
        'is_generic': getattr(task, 'is_generic', False),
        'backlog_id': task.backlog_id,
        'column_id': task.column_id,
        'sprint_id': task.sprint_id
    }

# --- Rotas de Frontend --- 

# GET /sprints/ - Página de Gerenciamento de Sprints
@sprints_bp.route('/', methods=['GET'])
def sprint_management_page():
    # Por enquanto, apenas renderiza o template. Os dados virão via API.
    return render_template('sprints/sprint_management.html', title="Gerenciamento de Sprints")

# --- API Endpoints para Sprints --- 

# GET /api/sprints - Listar todas as Sprints (VERSÃO OTIMIZADA)
@sprints_bp.route('/api/sprints', methods=['GET'])
def get_sprints():
    try:
        # Parâmetro para incluir sprints arquivadas (padrão: apenas ativas)
        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        
        # Busca sprints sem otimizações problemáticas (tasks é lazy='dynamic')
        if include_archived:
            sprints = Sprint.query.order_by(Sprint.start_date).all()
        else:
            sprints = Sprint.query.filter_by(is_archived=False).order_by(Sprint.start_date).all()
        
        # Usa o método to_dict otimizado do modelo
        sprints_data = []
        for sprint in sprints:
            try:
                sprint_data = sprint.to_dict()
                sprints_data.append(sprint_data)
            except Exception as sprint_error:
                current_app.logger.error(f"Erro ao processar sprint {sprint.id}: {sprint_error}")
                # Adiciona sprint com erro básico
                sprints_data.append({
                    'id': getattr(sprint, 'id', 'unknown'),
                    'name': f"Erro ao carregar sprint {getattr(sprint, 'id', '')}",
                    'start_date': None,
                    'end_date': None,
                    'goal': '',
                    'criticality': 'Normal',
                    'tasks': []
                })
        
        current_app.logger.info(f"Retornando {len(sprints_data)} sprints")
        return jsonify(sprints_data)
        
    except Exception as e:
        current_app.logger.error(f"Erro crítico ao buscar sprints: {str(e)}", exc_info=True)
        return jsonify({"message": f"Erro interno: {str(e)}"}), 500

# GET /api/sprints/<int:sprint_id> - Obter detalhes de uma Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>', methods=['GET'])
def get_sprint(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    return jsonify({
        'id': sprint.id,
        'name': sprint.name,
        'start_date': sprint.start_date.isoformat() if sprint.start_date else None,
        'end_date': sprint.end_date.isoformat() if sprint.end_date else None,
        'goal': sprint.goal,
        'criticality': sprint.criticality,
        # Adicionar lista de tasks e totais aqui futuramente
    })

# POST /api/sprints - Criar uma nova Sprint
@sprints_bp.route('/api/sprints', methods=['POST'])
def create_sprint():
    data = request.get_json()
    if not all(k in data for k in ('name', 'start_date', 'end_date')):
        abort(400, description="Campos 'name', 'start_date', e 'end_date' são obrigatórios.")

    try:
        start_date = datetime.fromisoformat(data['start_date'])
        end_date = datetime.fromisoformat(data['end_date'])
    except ValueError:
        abort(400, description="Formato de data inválido. Use ISO YYYY-MM-DD.")

    if start_date > end_date:
        abort(400, description="Data de início não pode ser posterior à data de fim.")
    
    new_sprint = Sprint(
        name=data['name'],
        start_date=start_date,
        end_date=end_date,
        goal=data.get('goal'),
        criticality=data.get('criticality', 'Normal') # Adiciona criticidade
    )
    db.session.add(new_sprint)
    db.session.commit()
    return jsonify(new_sprint.to_dict()), 201

# PUT /api/sprints/<int:sprint_id> - Atualizar uma Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>', methods=['PUT'])
def update_sprint(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    data = request.get_json()

    if 'name' in data:
        sprint.name = data['name']
    if 'start_date' in data:
        try:
            sprint.start_date = datetime.fromisoformat(data['start_date'])
        except ValueError:
            abort(400, "Formato de start_date inválido.")
    if 'end_date' in data:
        try:
            sprint.end_date = datetime.fromisoformat(data['end_date'])
        except ValueError:
            abort(400, "Formato de end_date inválido.")
    if 'goal' in data:
        sprint.goal = data.get('goal')
    if 'criticality' in data: # Adiciona atualização de criticidade
        sprint.criticality = data.get('criticality', sprint.criticality)


    if sprint.start_date and sprint.end_date and sprint.start_date > sprint.end_date:
        abort(400, description="Data de início não pode ser posterior à data de fim.")
        
    db.session.commit()
    return jsonify(sprint.to_dict())

# PUT /api/sprints/<int:sprint_id>/archive - Arquivar uma Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>/archive', methods=['PUT'])
def archive_sprint(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    
    if sprint.is_archived:
        return jsonify({'error': 'Sprint já está arquivada'}), 400
    
    # Arquiva a sprint
    sprint.is_archived = True
    sprint.archived_at = datetime.utcnow()
    sprint.archived_by = request.json.get('archived_by', 'Sistema') if request.json else 'Sistema'
    
    db.session.commit()
    return jsonify({
        'message': f'Sprint "{sprint.name}" arquivada com sucesso',
        'sprint': sprint.to_dict()
    })

# PUT /api/sprints/<int:sprint_id>/unarchive - Desarquivar uma Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>/unarchive', methods=['PUT'])
def unarchive_sprint(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    
    if not sprint.is_archived:
        return jsonify({'error': 'Sprint não está arquivada'}), 400
    
    # Desarquiva a sprint
    sprint.is_archived = False
    sprint.archived_at = None
    sprint.archived_by = None
    
    db.session.commit()
    return jsonify({
        'message': f'Sprint "{sprint.name}" desarquivada com sucesso',
        'sprint': sprint.to_dict()
    })

# GET /api/sprints/archived - Listar apenas Sprints Arquivadas
@sprints_bp.route('/api/sprints/archived', methods=['GET'])
def get_archived_sprints():
    try:
        sprints = Sprint.query.filter_by(is_archived=True).order_by(Sprint.archived_at.desc()).all()
        
        sprints_data = []
        for sprint in sprints:
            try:
                sprint_data = sprint.to_dict()
                sprints_data.append(sprint_data)
            except Exception as sprint_error:
                current_app.logger.error(f"Erro ao processar sprint arquivada {sprint.id}: {sprint_error}")
        
        return jsonify(sprints_data)
        
    except Exception as e:
        current_app.logger.error(f"Erro ao buscar sprints arquivadas: {str(e)}", exc_info=True)
        return jsonify({"message": f"Erro interno: {str(e)}"}), 500

# DELETE /api/sprints/<int:sprint_id> - Deletar uma Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>', methods=['DELETE'])
def delete_sprint(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    
    # Desassocia tarefas da sprint antes de excluí-la
    Task.query.filter_by(sprint_id=sprint_id).update({Task.sprint_id: None}, synchronize_session='fetch')
    
    db.session.delete(sprint)
    db.session.commit()
    # HTTP 204 No Content é mais apropriado para DELETE bem-sucedido sem corpo de resposta
    return '', 204

# --- Rotas Adicionais (Exemplo: Associar Tarefas) ---
# POST /api/sprints/<int:sprint_id>/tasks
# GET /api/sprints/<int:sprint_id>/tasks
# DELETE /api/sprints/<int:sprint_id>/tasks/<int:task_id>
# TODO: Implementar estas rotas na próxima fase. 

# --- Rotas API ---

# GET /sprints/api/generic-tasks - Lista todas as tarefas genéricas (SIMPLIFICADA)
@sprints_bp.route('/api/generic-tasks', methods=['GET'])
def get_generic_tasks_for_sprint_view():
    try:
        # VERSÃO SIMPLIFICADA: Query básica sem otimizações
        tasks = Task.query.filter(
            Task.is_generic == True, 
            Task.sprint_id == None
        ).order_by(Task.position).all()
        
        result = []
        for task in tasks:
            try:
                # Buscar informações do projeto de forma otimizada
                project_id = None
                project_name = 'Projeto não identificado'
                
                if task.backlog_id:
                    try:
                        from app.models import Backlog
                        backlog = Backlog.query.get(task.backlog_id)
                        if backlog and backlog.project_id:
                            project_id = backlog.project_id
                            # Simplificado - evita MacroService problemático
                            project_name = f'Projeto {project_id}'
                    except:
                        pass
                
                task_simple = {
                    'id': task.id,
                    'title': task.title or 'Tarefa sem título',
                    'description': task.description or '',
                    'priority': task.priority or 'Média',
                    'specialist_name': task.specialist_name or 'Não atribuído',
                    'estimated_effort': task.estimated_effort or 0,
                    'position': task.position or 0,
                    'status': str(task.status) if task.status else 'TODO',
                    'is_generic': True,
                    # Novas informações para o card
                    'project_id': project_id,
                    'project_name': project_name,
                    'backlog_id': task.backlog_id
                }
                result.append(task_simple)
            except Exception as task_error:
                current_app.logger.warning(f"Erro ao processar tarefa genérica {task.id}: {task_error}")
                continue
        
        current_app.logger.info(f"Retornando {len(result)} tarefas genéricas")
        return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"Erro ao buscar tarefas genéricas: {str(e)}", exc_info=True)
        return jsonify([]), 500

# POST /sprints/api/generic-tasks - Cria uma nova tarefa genérica
@sprints_bp.route('/api/generic-tasks', methods=['POST'])
def create_generic_task_from_sprint_view():
    data = request.get_json()
    if not data or not data.get('title'):
        abort(400, description="Título é obrigatório para tarefa genérica.")

    # Busca o backlog padrão
    default_backlog = Backlog.query.filter_by(project_id='default_project').first()
    if not default_backlog:
        abort(500, description="Backlog padrão não encontrado. Execute 'flask seed-db' primeiro.")

    # Busca a primeira coluna
    first_column = Column.query.order_by(Column.position).first()
    if not first_column:
        abort(500, description="Nenhuma coluna encontrada. Execute 'flask seed-db' primeiro.")

    # Calcula a posição da nova tarefa
    max_pos = db.session.query(db.func.max(Task.position)).filter(
        Task.backlog_id == default_backlog.id,
        Task.column_id == first_column.id
    ).scalar()
    new_position = (max_pos or -1) + 1

    new_task = Task(
        title=data['title'],
        description=data.get('description'),
        priority=data.get('priority', 'Média'),
        estimated_effort=data.get('estimated_hours'),
        specialist_name=data.get('specialist_name'),
        is_generic=True,
        status=TaskStatus.TODO,
        backlog_id=default_backlog.id,
        column_id=first_column.id,
        position=new_position
    )
    
    try:
        db.session.add(new_task)
        db.session.commit()
        return jsonify(serialize_task(new_task)), 201
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao criar tarefa genérica: {e}", exc_info=True)
        abort(500, description=f"Erro ao criar tarefa genérica: {str(e)}")

# PUT /sprints/api/generic-tasks/<task_id> - Atualiza uma tarefa genérica
@sprints_bp.route('/api/generic-tasks/<int:task_id>', methods=['PUT'])
def update_generic_task_from_sprint_view(task_id):
    task = Task.query.get_or_404(task_id)
    if not task.is_generic:
        abort(403, description="Esta rota é apenas para tarefas genéricas.")
    
    data = request.get_json()
    if not data:
        abort(400, description="Dados não fornecidos.")

    task.title = data.get('title', task.title)
    task.description = data.get('description', task.description)
    task.priority = data.get('priority', task.priority)
    task.estimated_effort = data.get('estimated_hours', task.estimated_effort)
    task.specialist_name = data.get('specialist_name', task.specialist_name)
    
    db.session.commit()
    return jsonify(serialize_task(task))

# DELETE /sprints/api/generic-tasks/<task_id> - Remove uma tarefa genérica
@sprints_bp.route('/api/generic-tasks/<int:task_id>', methods=['DELETE'])
def delete_generic_task_from_sprint_view(task_id):
    task = Task.query.get_or_404(task_id)
    if not task.is_generic:
        abort(403, description="Esta rota é apenas para tarefas genéricas.")

    # Se a tarefa estiver em uma sprint, precisa ser desassociada primeiro
    # ou a exclusão deve lidar com isso (ex: ON DELETE SET NULL no DB)
    # Por simplicidade, vamos permitir a exclusão direta.
    # O frontend deverá recarregar as sprints para refletir a remoção.
    
    db.session.delete(task)
    db.session.commit()
    return jsonify({"message": "Tarefa genérica excluída com sucesso."}), 200

# GET /sprints/report/<int:sprint_id> - Página de Relatório da Sprint
@sprints_bp.route('/report/<int:sprint_id>', methods=['GET'])
def sprint_report_page(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    
    # Agrupa tarefas por especialista
    specialist_tasks = defaultdict(list)
    specialist_effort = defaultdict(float)
    
    for task in sprint.tasks:
        if task.specialist_name:
            specialist_tasks[task.specialist_name].append(task)
            specialist_effort[task.specialist_name] += task.estimated_effort or 0
    
    return render_template(
        'sprints/sprint_report.html',
        sprint=sprint,
        specialist_tasks=dict(specialist_tasks),
        specialist_effort=dict(specialist_effort),
        title=f"Relatório - {sprint.name}"
    )

# GET /api/sprints/<int:sprint_id>/report - Obter dados do relatório da Sprint
@sprints_bp.route('/api/sprints/<int:sprint_id>/report', methods=['GET'])
def get_sprint_report(sprint_id):
    sprint = Sprint.query.get_or_404(sprint_id)
    
    # Agrupa tarefas por especialista
    specialist_tasks = defaultdict(list)
    specialist_effort = defaultdict(float)
    
    for task in sprint.tasks:
        if task.specialist_name:
            specialist_tasks[task.specialist_name].append(serialize_task(task))
            specialist_effort[task.specialist_name] += task.estimated_effort or 0
    
    return jsonify({
        'sprint': {
            'id': sprint.id,
            'name': sprint.name,
            'start_date': sprint.start_date.isoformat() if sprint.start_date else None,
            'end_date': sprint.end_date.isoformat() if sprint.end_date else None,
            'goal': sprint.goal,
            'criticality': sprint.criticality
        },
        'specialist_tasks': dict(specialist_tasks),
        'specialist_effort': dict(specialist_effort)
    })

# GET /sprints/consolidated-report - Página de seleção de Sprints para relatório
@sprints_bp.route('/consolidated-report', methods=['GET'])
def consolidated_report_page():
    # CORREÇÃO: Filtra apenas sprints ativas (não arquivadas)
    sprints = Sprint.query.filter(Sprint.is_archived != True).order_by(Sprint.start_date.desc()).all()
    return render_template(
        'sprints/consolidated_report_select.html',
        title="Relatório Consolidado de Sprints",
        sprints=sprints
    )

# POST /sprints/consolidated-report - Gerar relatório consolidado
@sprints_bp.route('/consolidated-report', methods=['POST'])
def generate_consolidated_report():
    sprint_ids = request.form.getlist('sprint_ids[]')
    if not sprint_ids:
        abort(400, description="Nenhuma sprint selecionada")

    # CORREÇÃO: Filtra apenas sprints ativas (não arquivadas) mesmo quando IDs específicos são fornecidos
    sprints = Sprint.query.filter(
        Sprint.id.in_(sprint_ids),
        Sprint.is_archived != True
    ).order_by(Sprint.start_date).all()
    if not sprints:
        abort(404, description="Nenhuma sprint ativa encontrada")

    # Cálculo de datas do período total
    start_date = min(sprint.start_date for sprint in sprints)
    end_date = max(sprint.end_date for sprint in sprints)

    # Contadores e acumuladores
    total_tasks = 0
    specialist_summary = defaultdict(lambda: {"name": "", "total_tasks": 0, "total_hours": 0})
    processed_sprints = []

    # Processa cada sprint
    for sprint in sprints:
        sprint_total_hours = 0
        processed_tasks = []

        # Busca todas as tarefas da sprint ordenadas por posição
        tasks = Task.query.filter_by(sprint_id=sprint.id).order_by(Task.position).all()
        
        for task in tasks:
            # Determina o status da tarefa baseado no nome da coluna
            status = "Em Andamento"
            if task.column_id:
                column = Column.query.get(task.column_id)
                if column and ('concluído' in column.name.lower() or 'concluido' in column.name.lower()):
                    status = "Concluído"
            
            # Processa a tarefa
            task_dict = {
                "id": task.id,
                "name": task.title,
                "specialist_name": task.specialist_name,
                "estimated_hours": task.estimated_effort or 0,
                "status": status
            }
            processed_tasks.append(task_dict)

            # Atualiza contadores
            total_tasks += 1
            sprint_total_hours += task.estimated_effort or 0

            # Atualiza resumo do especialista
            specialist = task.specialist_name or "Não Atribuído"
            specialist_summary[specialist]["name"] = specialist
            specialist_summary[specialist]["total_tasks"] += 1
            specialist_summary[specialist]["total_hours"] += task.estimated_effort or 0

        # Cria um dicionário com os dados processados da sprint
        sprint_dict = {
            "id": sprint.id,
            "name": sprint.name,
            "start_date": sprint.start_date,
            "end_date": sprint.end_date,
            "goal": sprint.goal,
            "criticality": sprint.criticality,
            "tasks": processed_tasks,
            "total_hours": sprint_total_hours
        }
        processed_sprints.append(sprint_dict)

    # Converte o resumo de especialistas para lista ordenada
    specialist_summary = sorted(
        specialist_summary.values(),
        key=lambda x: (-x["total_hours"], x["name"])
    )

    return render_template(
        'sprints/consolidated_report.html',
        sprints=processed_sprints,
        start_date=start_date,
        end_date=end_date,
        total_tasks=total_tasks,
        specialist_summary=specialist_summary,
        sprint_ids=sprint_ids  # Adicionando os IDs das sprints para o template
    )

# GET /sprints/export-consolidated-report - Exportar relatório consolidado para Excel
@sprints_bp.route('/export-consolidated-report', methods=['GET'])
def export_consolidated_report():
    sprint_ids = request.args.get('sprint_ids', '').split(',')
    if not sprint_ids or not sprint_ids[0]:
        abort(400, description="Nenhuma sprint selecionada")

    # CORREÇÃO: Filtra apenas sprints ativas (não arquivadas) mesmo quando IDs específicos são fornecidos
    sprints = Sprint.query.filter(
        Sprint.id.in_(sprint_ids),
        Sprint.is_archived != True
    ).order_by(Sprint.start_date).all()
    if not sprints:
        abort(404, description="Nenhuma sprint ativa encontrada")

    # Criar um novo workbook
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    # 1. Planilha de Resumo Geral
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Geral"
    
    # Calcular total de tarefas
    total_tasks = Task.query.filter(Task.sprint_id.in_(sprint_ids)).count()
    
    # Cabeçalhos
    headers = [
        ["Total de Sprints", str(len(sprints))],
        ["Período Início", min(sprint.start_date for sprint in sprints).strftime('%d/%m/%Y')],
        ["Período Fim", max(sprint.end_date for sprint in sprints).strftime('%d/%m/%Y')],
        ["Total de Tarefas", str(total_tasks)]
    ]
    
    for row_idx, (header, value) in enumerate(headers, 1):
        ws_resumo.cell(row=row_idx, column=1, value=header).font = header_font
        ws_resumo.cell(row=row_idx, column=2, value=value)
    
    # 2. Planilha de Capacidade por Especialista
    ws_especialistas = wb.create_sheet("Capacidade por Especialista")
    
    # Cabeçalhos
    headers = ["Especialista", "Total de Tarefas", "Horas Estimadas"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_especialistas.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Dados dos especialistas
    specialist_summary = defaultdict(lambda: {"name": "", "total_tasks": 0, "total_hours": 0})
    
    # Buscar todas as tarefas de uma vez
    tasks = Task.query.filter(Task.sprint_id.in_(sprint_ids)).all()
    
    for task in tasks:
        specialist = task.specialist_name or "Não Atribuído"
        specialist_summary[specialist]["name"] = specialist
        specialist_summary[specialist]["total_tasks"] += 1
        specialist_summary[specialist]["total_hours"] += task.estimated_effort or 0
    
    sorted_specialists = sorted(
        specialist_summary.values(),
        key=lambda x: (-x["total_hours"], x["name"])
    )
    
    for row_idx, specialist in enumerate(sorted_specialists, 2):
        ws_especialistas.cell(row=row_idx, column=1, value=specialist["name"])
        ws_especialistas.cell(row=row_idx, column=2, value=specialist["total_tasks"])
        ws_especialistas.cell(row=row_idx, column=3, value=round(specialist["total_hours"], 1))
    
    # 3. Planilha de Detalhes por Sprint
    ws_detalhes = wb.create_sheet("Detalhes por Sprint")
    
    # ✅ MELHORIA: Adicionando colunas do projeto (número e nome)
    headers = ["Sprint", "Período", "Tarefa", "Número Projeto", "Nome Projeto", "Especialista", "Horas Estimadas", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_detalhes.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Dados das tarefas
    current_row = 2
    for sprint in sprints:
        # Buscar tarefas da sprint ordenadas por posição
        sprint_tasks = Task.query.filter_by(sprint_id=sprint.id).order_by(Task.position).all()
        
        for task in sprint_tasks:
            status = "Em Andamento"
            if task.column_id:
                column = Column.query.get(task.column_id)
                if column and ('concluído' in column.name.lower() or 'concluido' in column.name.lower()):
                    status = "Concluído"
            
            # ✅ OBTENÇÃO DOS DADOS DO PROJETO
            project_number = "N/A"
            project_name = "Nome Indisponível"
            
            # Para tarefas genéricas, não há projeto associado
            if not getattr(task, 'is_generic', False) and task.backlog_id:
                try:
                    # Busca o backlog da tarefa
                    backlog = Backlog.query.get(task.backlog_id)
                    if backlog and backlog.project_id:
                        project_number = str(backlog.project_id)
                        
                        # Busca nome do projeto via MacroService (usando cache)
                        try:
                            from app.macro.services import MacroService
                            macro_service = MacroService()
                            project_details = macro_service.obter_detalhes_projeto(backlog.project_id)
                            if project_details:
                                # Busca o nome com fallback inteligente
                                project_name = (
                                    project_details.get('projeto') or 
                                    project_details.get('Projeto') or 
                                    project_details.get('cliente') or
                                    project_details.get('Cliente') or
                                    project_details.get('cliente_(completo)') or
                                    project_details.get('Cliente (Completo)') or
                                    f'Projeto {project_number}'
                                )
                        except Exception as e:
                            # Fallback silencioso
                            project_name = f'Projeto {project_number}'
                            current_app.logger.debug(f"Erro ao buscar nome do projeto {project_number}: {e}")
                        
                except Exception as e:
                    current_app.logger.debug(f"Erro ao obter dados do projeto para tarefa {task.id}: {e}")
            elif getattr(task, 'is_generic', False):
                project_number = "GENÉRICA"
                project_name = "Tarefa Genérica"
            
            # Escreve dados na planilha
            ws_detalhes.cell(row=current_row, column=1, value=sprint.name)
            ws_detalhes.cell(row=current_row, column=2, value=f"{sprint.start_date.strftime('%d/%m/%Y')} - {sprint.end_date.strftime('%d/%m/%Y')}")
            ws_detalhes.cell(row=current_row, column=3, value=task.title)
            ws_detalhes.cell(row=current_row, column=4, value=project_number)  # ✅ NOVA COLUNA
            ws_detalhes.cell(row=current_row, column=5, value=project_name)     # ✅ NOVA COLUNA
            ws_detalhes.cell(row=current_row, column=6, value=task.specialist_name or "Não Atribuído")
            ws_detalhes.cell(row=current_row, column=7, value=round(task.estimated_effort or 0, 1))
            ws_detalhes.cell(row=current_row, column=8, value=status)
            current_row += 1
    
    # ✅ OTIMIZAÇÃO: Ajustar larguras das colunas com configuração específica
    # Ajusta larguras automáticas para ws_resumo e ws_especialistas
    for ws in [ws_resumo, ws_especialistas]:
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # ✅ CONFIGURAÇÃO ESPECÍFICA para ws_detalhes (com as novas colunas)
    column_widths = {
        'A': 20,  # Sprint
        'B': 22,  # Período
        'C': 35,  # Tarefa
        'D': 15,  # Número Projeto
        'E': 30,  # Nome Projeto
        'F': 18,  # Especialista
        'G': 15,  # Horas Estimadas
        'H': 12   # Status
    }
    
    for column_letter, width in column_widths.items():
        ws_detalhes.column_dimensions[column_letter].width = width
    
    # Salvar o arquivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"relatorio_consolidado_sprints_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# POST /api/sprints/tasks/<int:task_id>/move-to-backlog - Mover tarefa para o backlog (ou genéricas)
@sprints_bp.route('/api/sprints/tasks/<int:task_id>/move-to-backlog', methods=['POST'])
def move_task_to_backlog(task_id):
    current_app.logger.info(f">>> ROTA ACESSADA (move-to-backlog) para task {task_id} com método {request.method}")
    task = Task.query.get_or_404(task_id)
    current_app.logger.info(f"Detalhes da tarefa antes da alteração: ID={task.id}, SprintID={task.sprint_id}, is_generic={task.is_generic}")

    if task.sprint_id is None:
        current_app.logger.warn(f"Tarefa {task_id} já não está em uma sprint.")
        # Ainda assim, retorna sucesso pois o estado desejado (fora da sprint) é alcançado.
        # O frontend vai recarregar as listas, então a tarefa aparecerá no lugar certo.
        return jsonify(serialize_task(task)), 200

    try:
        task.sprint_id = None
        # A posição da tarefa no backlog ou na lista de genéricas não é alterada aqui,
        # ela mantém sua última posição conhecida nessas listas.
        # Se for necessário re-posicionar ao final da lista de backlog/genéricas,
        # seria preciso adicionar lógica aqui ou no frontend para recalcular.
        # Por ora, simplificamos e a tarefa apenas "volta" para onde estava.

        db.session.commit()
        current_app.logger.info(f"Tarefa {task_id} desvinculada da sprint. Novo SprintID: {task.sprint_id}")
        
        # Serializa e retorna a tarefa atualizada.
        # O frontend usará o campo 'is_generic' para decidir qual lista recarregar.
        return jsonify(serialize_task(task)), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao mover tarefa {task_id} para fora da sprint: {e}", exc_info=True)
        return jsonify({"message": "Erro interno ao tentar mover a tarefa."}), 500

# --- NOVAS ROTAS PARA CÁLCULO AUTOMÁTICO DE DATAS ---

@sprints_bp.route('/api/sprints/<int:sprint_id>/calculate-dates', methods=['POST'])
def calculate_sprint_dates(sprint_id):
    """API para calcular datas sequenciais das tarefas de uma sprint"""
    try:
        from ..utils.base_service import DateCalculationService
        
        sprint = Sprint.query.get_or_404(sprint_id)
        
        if not sprint.start_date:
            return jsonify({
                'success': False, 
                'error': 'Sprint precisa ter data de início configurada'
            }), 400
        
        # Busca tarefas da sprint ordenadas por posição
        tasks = Task.query.filter_by(sprint_id=sprint_id).order_by(Task.position).all()
        
        if not tasks:
            return jsonify({
                'success': True,
                'message': 'Nenhuma tarefa encontrada na sprint',
                'updated_tasks': []
            })
        
        # Converte tarefas para formato do serviço
        task_data = []
        for task in tasks:
            task_data.append({
                'id': task.id,
                'specialist_name': task.specialist_name,
                'estimated_effort': task.estimated_effort or 0,
                'title': task.title
            })
        
        # Calcula datas sequenciais
        updated_task_data = DateCalculationService.calculate_sequential_dates(
            task_data, sprint.start_date
        )
        
        # Atualiza tarefas no banco
        updated_tasks = []
        for task_info in updated_task_data:
            task = Task.query.get(task_info['id'])
            if task:
                task.start_date = datetime.fromisoformat(task_info['start_date']) if task_info.get('start_date') else None
                task.due_date = datetime.fromisoformat(task_info['due_date']) if task_info.get('due_date') else None
                task.updated_at = datetime.utcnow()
                updated_tasks.append(serialize_task(task))
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Datas calculadas para {len(updated_tasks)} tarefas',
            'updated_tasks': updated_tasks
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao calcular datas da sprint {sprint_id}: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Erro ao calcular datas: {str(e)}'
        }), 500

@sprints_bp.route('/api/sprints/<int:sprint_id>/capacity-alerts', methods=['GET'])
def get_sprint_capacity_alerts(sprint_id):
    """API para obter alertas e sugestões de capacidade da sprint"""
    try:
        from ..utils.base_service import DateCalculationService
        
        sprint = Sprint.query.get_or_404(sprint_id)
        
        if not sprint.start_date or not sprint.end_date:
            return jsonify({
                'success': False,
                'error': 'Sprint precisa ter datas de início e fim configuradas'
            }), 400
        
        # Busca tarefas da sprint
        tasks = Task.query.filter_by(sprint_id=sprint_id).all()
        
        # Converte para formato do serviço
        task_data = []
        for task in tasks:
            task_data.append({
                'id': task.id,
                'specialist_name': task.specialist_name,
                'estimated_effort': task.estimated_effort or 0,
                'title': task.title
            })
        
        # Datas da sprint
        sprint_dates = {
            'start_date': sprint.start_date,
            'end_date': sprint.end_date
        }
        
        # Calcula alertas de capacidade
        alerts = DateCalculationService.calculate_sprint_capacity_alerts(
            task_data, sprint_dates
        )
        
        return jsonify({
            'success': True,
            'sprint_id': sprint_id,
            'sprint_name': sprint.name,
            'alerts': alerts
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao calcular alertas da sprint {sprint_id}: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Erro ao calcular alertas: {str(e)}'
        }), 500

@sprints_bp.route('/api/sprints/batch-calculate-dates', methods=['POST'])
def batch_calculate_sprint_dates():
    """API para calcular datas de múltiplas sprints em lote"""
    try:
        data = request.get_json()
        sprint_ids = data.get('sprint_ids', [])
        
        if not sprint_ids:
            return jsonify({
                'success': False,
                'error': 'Lista de IDs de sprints é obrigatória'
            }), 400
        
        results = []
        total_updated = 0
        
        for sprint_id in sprint_ids:
            try:
                sprint = Sprint.query.get(sprint_id)
                if not sprint or not sprint.start_date:
                    results.append({
                        'sprint_id': sprint_id,
                        'success': False,
                        'error': 'Sprint não encontrada ou sem data de início'
                    })
                    continue
                
                # Reutiliza a lógica da rota individual
                response = calculate_sprint_dates(sprint_id)
                response_data = response.get_json()
                
                if response_data.get('success'):
                    total_updated += len(response_data.get('updated_tasks', []))
                    results.append({
                        'sprint_id': sprint_id,
                        'sprint_name': sprint.name,
                        'success': True,
                        'updated_count': len(response_data.get('updated_tasks', []))
                    })
                else:
                    results.append({
                        'sprint_id': sprint_id,
                        'success': False,
                        'error': response_data.get('error', 'Erro desconhecido')
                    })
                    
            except Exception as e:
                results.append({
                    'sprint_id': sprint_id,
                    'success': False,
                    'error': str(e)
                })
        
        return jsonify({
            'success': True,
            'message': f'Processamento concluído. {total_updated} tarefas atualizadas.',
            'results': results,
            'total_updated_tasks': total_updated
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro no cálculo em lote: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Erro no processamento em lote: {str(e)}'
        }), 500

@sprints_bp.route('/api/sprints/<int:sprint_id>/sync-backlog-dates', methods=['POST'])
def sync_backlog_dates(sprint_id):
    """API para sincronizar datas calculadas de volta para o backlog"""
    try:
        sprint = Sprint.query.get_or_404(sprint_id)
        
        # Busca tarefas da sprint que tenham datas calculadas
        tasks = Task.query.filter(
            Task.sprint_id == sprint_id,
            Task.start_date.isnot(None),
            Task.due_date.isnot(None)
        ).all()
        
        if not tasks:
            return jsonify({
                'success': True,
                'message': 'Nenhuma tarefa com datas para sincronizar',
                'synced_count': 0
            })
        
        # Sincroniza datas para o backlog (mantém as datas mesmo quando movida de volta)
        synced_count = 0
        for task in tasks:
            # As datas já estão salvas no banco, não precisamos fazer nada específico
            # O backlog vai mostrar essas datas automaticamente
            synced_count += 1
        
        return jsonify({
            'success': True,
            'message': f'Datas sincronizadas para {synced_count} tarefas',
            'synced_count': synced_count,
            'sprint_name': sprint.name
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao sincronizar datas da sprint {sprint_id}: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Erro na sincronização: {str(e)}'
        }), 500

# --- FIM NOVAS ROTAS --- 

# Nova rota para clonagem de tarefas
@sprints_bp.route('/api/sprints/tasks/<int:task_id>/clone', methods=['POST'])
def clone_task(task_id):
    """
    Clona uma tarefa da sprint para o backlog.
    Se há sprint ativa, o clone vai direto para ela.
    Caso contrário, fica no backlog do projeto.
    """
    try:
        # Busca a tarefa original
        original_task = Task.query.get_or_404(task_id)
        current_app.logger.info(f"Clonando tarefa {task_id}: {original_task.title}")
        
        # Verifica se a tarefa está em uma sprint
        if not original_task.sprint_id:
            return jsonify({'error': 'Só é possível clonar tarefas que estão em uma sprint'}), 400
        
        # Busca a sprint original (mesma sprint da tarefa original)
        original_sprint = Sprint.query.get(original_task.sprint_id)
        if not original_sprint:
            return jsonify({'error': 'Sprint da tarefa original não encontrada'}), 400
        
        # Cria novo título com sufixo
        clone_title = f"{original_task.title} (Clone)"
        
        # Cria a tarefa clonada
        cloned_task = Task(
            title=clone_title,
            description=original_task.description,
            priority=original_task.priority,
            estimated_effort=original_task.estimated_effort,
            specialist_name=original_task.specialist_name,
            backlog_id=original_task.backlog_id,  # Mantém no mesmo projeto/backlog
            column_id=original_task.column_id,    # Mantém na mesma coluna
            status=TaskStatus.TODO,               # Status inicial sempre TODO
            is_generic=False,                     # Clone não é genérico (vai para rastreamento)
            sprint_id=original_task.sprint_id     # ✅ MESMA SPRINT DA TAREFA ORIGINAL
        )
        
        # Calcula próxima posição na mesma sprint
        max_position = db.session.query(db.func.max(Task.position)).filter_by(sprint_id=original_task.sprint_id).scalar()
        cloned_task.position = (max_position or 0) + 1
        current_app.logger.info(f"Clone adicionado à mesma sprint da original: {original_sprint.name}")
        
        # Salva no banco
        db.session.add(cloned_task)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Tarefa clonada com sucesso',
            'original_task': serialize_task(original_task),
            'cloned_task': serialize_task(cloned_task),
            'added_to_sprint': original_sprint.name
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao clonar tarefa {task_id}: {str(e)}", exc_info=True)
        return jsonify({'error': f'Erro ao clonar tarefa: {str(e)}'}), 500